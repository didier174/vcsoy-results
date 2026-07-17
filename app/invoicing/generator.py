"""
Génération de la facture, dans les deux formats demandés :
- Excel : à partir du modèle fourni (assets/invoice_template.xlsx), en ne
  modifiant que les cellules variables (dates, montants, adresse du
  participant...) — toute la mise en forme, le logo et les informations de
  l'émetteur (CA2D) restent ceux du modèle d'origine.
- PDF : mise en page indépendante, générée avec reportlab (pas de
  dépendance à un logiciel externe type LibreOffice, ce qui garde le
  déploiement simple).
"""

import io
import os
from copy import copy

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import range_boundaries, get_column_letter
from PIL import Image as PILImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "invoice_template.xlsx")
IMG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "static", "img")

# Mappage des lignes, répliquant la structure du modèle fourni : le
# produit VCSOY occupe toujours les lignes 27-30 (intitulé, ligne 27,
# portant la quantité/le prix de l'ensemble, puis 3 puces descriptives
# décalées à droite, lignes 28-30, sans montant propre).
ROW_VCSOY_HEADING = 27
ROW_VCSOY_PLAIN_BULLETS = [28, 29, 30]

# Les produits du catalogue (titre + jusqu'à 3 puces chacun, nombre de
# lignes donc variable) prennent ensuite les lignes déjà formatées dans le
# modèle (31, 33, 34, 35 — la ligne 32 sert d'espacement). Si plus de
# lignes sont nécessaires, on en insère juste avant la ligne d'espacement
# suivante (36), en reproduisant le style de la dernière ligne du modèle
# (voir _copy_row_style) — le bloc sous-total/TPS/TVQ/total et tout ce qui
# suit (mention d'exportation, coordonnées de paiement) est alors décalé
# d'autant de lignes.
CATALOG_ROWS_AVAILABLE = [31, 33, 34, 35]
CATALOG_BLOCK_LAST_ROW = 35
TOTALS_FIRST_ROW = 38  # SOUS-TOTAL ; TPS/TVQ/TOTAL CAD suivent sur les 3 lignes suivantes
PAYMENT_INTRO_ROW = 46

# Décalage visuel (indentation Excel native) des puces descriptives d'un
# produit (VCSOY ou catalogue), pour bien montrer qu'elles font partie de
# la ligne titre au-dessus plutôt que d'être des produits distincts.
BULLET_INDENT = 2

# Taille cible du logo dans le classeur Excel (en pixels) et dans le PDF
# (en points) — la largeur est recalculée pour chaque logo afin de
# préserver son ratio d'origine et éviter toute déformation.
XLSX_LOGO_HEIGHT_PX = 95
PDF_LOGO_HEIGHT_PT = 62

# Logo (français/anglais) à utiliser par édition. Le logo anglais 2028
# n'a pas encore été fourni : on utilise en attendant le logo français de
# la même édition, pour ne jamais afficher un logo d'une mauvaise édition.
LOGO_FILES = {
    ("blanche", "fr"): "logo_annee_fr.png",
    ("blanche", "en"): "logo_annee_fr.png",  # même logo qu'en français (demande explicite)
    ("2027", "fr"): "logo_2027_fr.png",
    ("2027", "en"): "logo_2027_en.png",
    ("2028", "fr"): "logo_2028_fr.png",
    ("2028", "en"): "logo_2028_en.png",  # à fournir ; repli automatique sinon
    ("2029", "fr"): "logo_2029_fr.png",
    ("2029", "en"): "logo_2029_en.png",
    ("2030", "fr"): "logo_2030_fr.png",
    ("2030", "en"): "logo_2030_en.png",
}


def resolve_logo_path(edition_id, language):
    """Retourne le chemin du logo à utiliser sur la facture, avec repli sûr."""
    filename = LOGO_FILES.get((edition_id, language))
    if filename:
        path = os.path.join(IMG_DIR, filename)
        if os.path.exists(path):
            return path

    # Repli 1 : logo français de la même édition (toujours fourni pour l'instant)
    fallback_filename = LOGO_FILES.get((edition_id, "fr"))
    if fallback_filename:
        path = os.path.join(IMG_DIR, fallback_filename)
        if os.path.exists(path):
            return path

    # Repli 2 : logo français 2027 (garantit qu'il y a toujours un logo affiché)
    path = os.path.join(IMG_DIR, "logo_2027_fr.png")
    return path if os.path.exists(path) else None


def _logo_pixel_size(logo_path, target_height):
    with PILImage.open(logo_path) as pil_img:
        ratio = pil_img.width / pil_img.height
    height = target_height
    width = int(height * ratio)
    return width, height

LABELS = {
    "fr": {
        "invoice_title": "FACTURE",
        "date": "DATE",
        "invoice_no": "No de facture :",
        "customer_no": "No de client :",
        "bill_to": "Facturé à",
        "description": "DESCRIPTION",
        "quantity": "QUANTITÉ",
        "unit_price": "PRIX UNITAIRE",
        "total": "TOTAL",
        "subtotal": "SOUS-TOTAL",
        "gst": "TPS",
        "qst": "TVQ",
        "total_cad": "TOTAL CAD",
        "export_note": "TPS/TVQ : 0 % – Exportation de services à un non-résident (fourniture détaxée)",
        "payment_intro": "Virement bancaire ou chèque",
    },
    "en": {
        "invoice_title": "INVOICE",
        "date": "DATE",
        "invoice_no": "Invoice No:",
        "customer_no": "Customer No:",
        "bill_to": "Bill to",
        "description": "DESCRIPTION",
        "quantity": "QUANTITY",
        "unit_price": "UNIT PRICE",
        "total": "TOTAL",
        "subtotal": "SUB TOTAL",
        "gst": "TPS",
        "qst": "TVQ",
        "total_cad": "TOTAL CAD",
        "export_note": "GST/QST: 0% – Export of services to a non-resident (zero-rated supply)",
        "payment_intro": "Wire transfert or check",
    },
}


def _labels_for(language):
    return LABELS.get(language, LABELS["fr"])


def _bill_to_address_lines(invoice):
    addr_line = invoice.bill_to_address1 or ""
    if invoice.bill_to_address2:
        addr_line = f"{addr_line}, {invoice.bill_to_address2}" if addr_line else invoice.bill_to_address2
    city_line = ", ".join(p for p in [invoice.bill_to_city, invoice.bill_to_postal_code] if p)
    return addr_line, city_line


# ------------------------------------------------------------------ Excel

def _shift_merges_and_insert_rows(ws, insert_at, count):
    """Insère `count` lignes à la position `insert_at`. openpyxl décale
    correctement les valeurs/styles des cellules déjà en place, mais PAS
    les plages de cellules fusionnées situées à/après ce point : on les
    démérge avant l'insertion puis on les refusionne à leur nouvelle
    position (+count), sans quoi elles restent visuellement à l'ancien
    endroit alors que le contenu, lui, a bien été décalé."""
    if count <= 0:
        return
    to_shift = [str(mc) for mc in ws.merged_cells.ranges if mc.min_row >= insert_at]
    for rng in to_shift:
        ws.unmerge_cells(rng)

    ws.insert_rows(insert_at, amount=count)

    for rng in to_shift:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        new_rng = f"{get_column_letter(min_col)}{min_row + count}:{get_column_letter(max_col)}{max_row + count}"
        ws.merge_cells(new_rng)


def _copy_row_style(ws, source_row, target_row, min_col=1, max_col=36):
    """Reproduit la mise en forme (police, bordures, alignement, format de
    nombre, hauteur) d'une ligne déjà présente dans le modèle sur une
    ligne nouvellement insérée (colonnes A à AJ = 1 à 36), qui sinon reste
    sans aucun style (police/taille par défaut, pas de bordure)."""
    for col in range(min_col, max_col + 1):
        src = ws.cell(row=source_row, column=col)
        dst = ws.cell(row=target_row, column=col)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def _merge_item_row(ws, row):
    """Reproduit sur une ligne nouvellement insérée le même motif de
    fusion de cellules que les autres lignes d'article du modèle (colonnes
    description / quantité / prix unitaire / total) : `insert_rows` ne crée
    aucune fusion pour les lignes qu'il ajoute, contrairement à celles
    qu'il décale (voir _shift_merges_and_insert_rows). Sans cela, la
    quantité/le total d'une ligne insérée se retrouve dans une seule
    cellule bien trop étroite pour l'afficher, d'où des "###" à l'ouverture."""
    ws.merge_cells(f"A{row}:T{row}")
    ws.merge_cells(f"U{row}:Y{row}")
    ws.merge_cells(f"Z{row}:AE{row}")
    ws.merge_cells(f"AF{row}:AJ{row}")


def fill_invoice_xlsx(invoice):
    """Retourne un BytesIO contenant le fichier Excel de la facture."""
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Feuil1"]
    labels = _labels_for(invoice.language)

    # Logo de l'édition en cours, dans la langue de la facture. On ne
    # touche à aucune fusion de cellule ni à aucune autre mise en forme du
    # modèle : seule l'image est remplacée, à la même position (B1).
    logo_path = resolve_logo_path(invoice.edition_id, invoice.language)
    ws._images = []
    if logo_path:
        width, height = _logo_pixel_size(logo_path, XLSX_LOGO_HEIGHT_PX)
        xl_img = XLImage(logo_path)
        xl_img.width = width
        xl_img.height = height
        ws.add_image(xl_img, "B1")

    ws["AH1"] = labels["invoice_title"]
    ws["W5"] = labels["invoice_no"]
    ws["W6"] = labels["customer_no"]
    ws["W10"] = labels["bill_to"]
    ws["A22"] = labels["description"]
    ws["U22"] = labels["quantity"]
    ws["Z22"] = labels["unit_price"]
    ws["AF22"] = labels["total"]

    ws["AC4"] = invoice.invoice_date
    ws["AC5"] = invoice.invoice_number
    ws["AC6"] = invoice.customer_number

    ws["W12"] = invoice.bill_to_contact_name
    ws["W14"] = invoice.bill_to_company_name
    addr_line, city_line = _bill_to_address_lines(invoice)
    ws["W15"] = addr_line
    ws["W16"] = city_line
    ws["W17"] = invoice.bill_to_country

    def _write_description(row, text, bold, indent=0):
        cell = ws[f"A{row}"]
        cell.value = text
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=bold)
        # Toujours réécrire l'indentation (même à 0) : sans ça, une ligne
        # sans indent hérite silencieusement de l'indentation résiduelle du
        # modèle Excel d'origine à cette position (ex. Goodies aligné comme
        # une puce du produit VCSOY plutôt qu'au début de sa propre ligne).
        existing = cell.alignment
        cell.alignment = Alignment(
            horizontal=existing.horizontal, vertical=existing.vertical,
            wrap_text=existing.wrap_text, indent=indent,
        )

    def _write_amounts(row, quantity, unit_price, total):
        qty_cell = ws[f"U{row}"]
        qty_cell.value = quantity
        # Toujours un nombre entier (pas de décimales) : une quantité est
        # plafonnée à 3 chiffres (voir validation du formulaire) et un
        # format avec décimales/séparateur de milliers prend trop de place
        # dans la cellule, ce qui affiche "###" à l'ouverture du fichier.
        qty_cell.number_format = "0"
        ws[f"Z{row}"] = unit_price
        ws[f"AF{row}"] = total

    def _clear_row(row):
        _write_description(row, "", bold=False)
        _write_amounts(row, None, None, None)

    # Le produit VCSOY (s'il est sélectionné) occupe toujours les lignes
    # 27-30 ; sinon on efface les lignes d'exemple du modèle d'origine.
    vcsoy_heading_item = next((i for i in invoice.line_items if i.get("role") == "vcsoy_heading"), None)
    vcsoy_plain_items = [i for i in invoice.line_items if i.get("role") == "vcsoy_plain"]
    other_items = [i for i in invoice.line_items if i.get("role") not in ("vcsoy_heading", "vcsoy_plain")]

    if vcsoy_heading_item:
        _write_description(ROW_VCSOY_HEADING, vcsoy_heading_item.get("description", ""), bold=True)
        _write_amounts(
            ROW_VCSOY_HEADING, vcsoy_heading_item.get("quantity", 0),
            vcsoy_heading_item.get("unit_price", 0), vcsoy_heading_item.get("total", 0),
        )
        plain_bullet_rows = iter(ROW_VCSOY_PLAIN_BULLETS)
        for item in vcsoy_plain_items:
            row = next(plain_bullet_rows, None)
            if row is not None:
                _write_description(row, item.get("description", ""), bold=False, indent=BULLET_INDENT)
                _write_amounts(row, None, None, None)
        for row in plain_bullet_rows:
            _clear_row(row)
    else:
        for row in [ROW_VCSOY_HEADING] + ROW_VCSOY_PLAIN_BULLETS:
            _clear_row(row)

    # Produits du catalogue (titre + jusqu'à 3 puces chacun, nombre de
    # lignes variable) : utilise les lignes déjà formatées du modèle, et en
    # insère si besoin (voir _shift_merges_and_insert_rows/_copy_row_style),
    # ce qui décale d'autant le bloc sous-total et tout ce qui suit.
    needed = len(other_items)
    available = len(CATALOG_ROWS_AVAILABLE)
    extra_needed = max(0, needed - available)

    if extra_needed:
        insert_at = CATALOG_BLOCK_LAST_ROW + 1
        _shift_merges_and_insert_rows(ws, insert_at, extra_needed)
        for offset in range(extra_needed):
            _copy_row_style(ws, CATALOG_BLOCK_LAST_ROW, insert_at + offset)
            _merge_item_row(ws, insert_at + offset)
        catalog_rows = CATALOG_ROWS_AVAILABLE + list(range(insert_at, insert_at + extra_needed))
    else:
        catalog_rows = list(CATALOG_ROWS_AVAILABLE)

    catalog_rows_iter = iter(catalog_rows)
    for item in other_items:
        row = next(catalog_rows_iter, None)
        if row is None:
            break
        role = item.get("role")
        if role == "catalog_bullet":
            _write_description(row, item.get("description", ""), bold=False, indent=BULLET_INDENT)
            _write_amounts(row, None, None, None)
        else:
            # "catalog_heading", ou "standalone"/générique (anciennes factures).
            _write_description(row, item.get("description", ""), bold=False)
            _write_amounts(row, item.get("quantity", 0), item.get("unit_price", 0), item.get("total", 0))
    # Lignes du bloc catalogue non utilisées : les blanchir (au cas où le
    # modèle y contiendrait un jour un exemple, comme pour le bloc VCSOY).
    for row in catalog_rows_iter:
        _clear_row(row)

    totals_offset = extra_needed
    subtotal_row = TOTALS_FIRST_ROW + totals_offset
    gst_row = subtotal_row + 1
    qst_row = subtotal_row + 2
    total_row = subtotal_row + 3
    payment_intro_row = PAYMENT_INTRO_ROW + totals_offset

    ws[f"Y{subtotal_row}"] = labels["subtotal"]
    ws[f"Y{gst_row}"] = labels["gst"]
    ws[f"Y{qst_row}"] = labels["qst"]
    ws[f"Y{total_row}"] = labels["total_cad"]
    ws[f"AF{subtotal_row}"] = invoice.subtotal
    ws[f"AF{gst_row}"] = invoice.gst_amount
    ws[f"AF{qst_row}"] = invoice.qst_amount
    ws[f"AF{total_row}"] = invoice.total_amount
    ws[f"B{payment_intro_row}"] = labels["payment_intro"]

    # La mention d'exportation est encadrée d'un cadre en pointillés qui
    # s'arrête à la colonne T (voir le modèle) : on fusionne B:T sur la
    # ligne TVQ et on active le retour à la ligne automatique pour que le
    # texte ne déborde jamais au-delà de ce cadre, quelle que soit la langue.
    if invoice.is_export:
        ws[f"B{qst_row}"] = labels["export_note"]
        ws.merge_cells(f"B{qst_row}:T{qst_row}")
        note_cell = ws[f"B{qst_row}"]
        note_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
    else:
        ws[f"B{qst_row}"] = ""

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# -------------------------------------------------------------------- PDF

def render_invoice_pdf(invoice):
    labels = _labels_for(invoice.language)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter, topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    normal = ParagraphStyle("normal9", parent=styles["Normal"], fontSize=9, leading=13)
    title_style = ParagraphStyle("title", parent=styles["Title"], alignment=TA_RIGHT, fontSize=24)
    note_style = ParagraphStyle("note", parent=normal, fontSize=8, textColor=colors.grey)
    footer_style = ParagraphStyle("footer", parent=normal, fontSize=8)

    elements = []

    logo_path = resolve_logo_path(invoice.edition_id, invoice.language)
    company_para = Paragraph(
        "<b>CA2D Inc.</b><br/>1203 Avenue Bernard<br/>MONTREAL, QC, H2V 1V7<br/>CANADA<br/>"
        "Phone: +1 514 690 3652<br/>TPS: 811652985RT001<br/>TVQ: 1222850718",
        normal,
    )
    if logo_path:
        width, height = _logo_pixel_size(logo_path, PDF_LOGO_HEIGHT_PT)
        logo_flowable = RLImage(logo_path, width=width, height=height)
        left_column = [logo_flowable, Spacer(1, 8), company_para]
    else:
        left_column = [company_para]

    header_data = [[left_column, Paragraph(labels["invoice_title"], title_style)]]
    header_table = Table(header_data, colWidths=[300, 195])
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(header_table)
    elements.append(Spacer(1, 16))

    meta_data = [
        [labels["date"], invoice.invoice_date.strftime("%Y-%m-%d") if invoice.invoice_date else ""],
        [labels["invoice_no"], invoice.invoice_number or ""],
        [labels["customer_no"], invoice.customer_number or ""],
    ]
    meta_table = Table(meta_data, colWidths=[85, 110])
    meta_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "LEFT"),
    ]))

    addr_line, city_line = _bill_to_address_lines(invoice)
    bill_to_parts = [
        invoice.bill_to_contact_name, invoice.bill_to_company_name, addr_line, city_line, invoice.bill_to_country,
    ]
    bill_to_text = "<br/>".join(p for p in bill_to_parts if p)
    bill_to_para = Paragraph(f"<b>{labels['bill_to']}</b><br/>{bill_to_text}", normal)

    top_table = Table([[bill_to_para, meta_table]], colWidths=[300, 195])
    top_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    elements.append(top_table)
    elements.append(Spacer(1, 22))

    table_data = [[labels["description"], labels["quantity"], labels["unit_price"], labels["total"]]]
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
        ("LEFTPADDING", (0, 1), (0, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for i, item in enumerate(invoice.line_items, start=1):
        role = item.get("role")
        # .get(..., 0) plutôt que item[...] : certaines factures créées
        # avec une version antérieure du code n'ont pas ces clés sur toutes
        # les lignes (voir historique) — éviter un KeyError qui empêcherait
        # de télécharger une facture existante.
        if role == "vcsoy_heading":
            # Intitulé du produit VCSOY, portant sa quantité/prix/total.
            table_data.append([
                item.get("description", ""),
                f"{item.get('quantity', 0):.0f}",
                f"{item.get('unit_price', 0):,.2f} $",
                f"{item.get('total', 0):,.2f} $",
            ])
            style_commands.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))
        elif role == "vcsoy_plain":
            # Puce descriptive du produit VCSOY, sans prix propre : décalée
            # à droite pour bien montrer qu'elle fait partie de l'intitulé
            # ci-dessus plutôt que d'être un produit distinct.
            table_data.append(["- " + item.get("description", ""), "", "", ""])
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))
            style_commands.append(("LEFTPADDING", (0, i), (0, i), 24))
        elif role == "catalog_heading":
            # Produit du catalogue : même présentation que le VCSOY (titre
            # portant quantité/prix/total, éventuellement suivi de puces).
            table_data.append([
                item.get("description", ""),
                f"{item.get('quantity', 0):.0f}",
                f"{item.get('unit_price', 0):,.2f} $",
                f"{item.get('total', 0):,.2f} $",
            ])
            style_commands.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))
        elif role == "catalog_bullet":
            # Puce descriptive d'un produit du catalogue, sans prix propre.
            table_data.append(["- " + item.get("description", ""), "", "", ""])
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))
            style_commands.append(("LEFTPADDING", (0, i), (0, i), 24))
        elif "total" in item:
            # Produit indépendant (Right to use trademark, Goodies, ...) :
            # une seule ligne, mise en gras comme l'intitulé du VCSOY, sans
            # tiret puisqu'il n'y a pas de puces de détail en dessous.
            table_data.append([
                item.get("description", ""),
                f"{item.get('quantity', 0):.0f}",
                f"{item.get('unit_price', 0):,.2f} $",
                f"{item.get('total', 0):,.2f} $",
            ])
            style_commands.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))
        else:
            # Ligne descriptive sans prix propre (cas générique restant).
            table_data.append(["- " + item.get("description", ""), "", "", ""])
            style_commands.append(("ALIGN", (0, i), (0, i), "LEFT"))

    products_table = Table(table_data, colWidths=[255, 70, 90, 90])
    products_table.setStyle(TableStyle(style_commands))
    elements.append(products_table)
    elements.append(Spacer(1, 16))

    totals_data = [
        [labels["subtotal"], f"{invoice.subtotal:,.2f} $"],
        [labels["gst"], f"{invoice.gst_amount:,.2f} $"],
        [labels["qst"], f"{invoice.qst_amount:,.2f} $"],
        [labels["total_cad"], f"{invoice.total_amount:,.2f} $"],
    ]
    totals_table = Table(totals_data, colWidths=[110, 100])
    totals_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    wrapper = Table([[None, totals_table]], colWidths=[295, 210])
    elements.append(wrapper)

    if invoice.is_export:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(labels["export_note"], note_style))

    elements.append(Spacer(1, 34))
    elements.append(Paragraph(
        f"<b>{labels['payment_intro']}</b><br/>CA2D Inc<br/>"
        "Bank: RBC 1 Place Ville-Marie, Montréal, QC H3B 3Y1<br/>"
        "Swift: ROYCCAT2MIC<br/>"
        "Institution Number: 003 &nbsp;&nbsp; Transit Number: 04896 &nbsp;&nbsp; Account Number: 1009836",
        footer_style,
    ))

    doc.build(elements)
    buf.seek(0)
    return buf
