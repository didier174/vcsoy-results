"""
Validation du fichier Excel de résultats de tests mystères, selon les
règles définies dans le cahier des charges :

- Colonne A (ID Mystery Test) : chaîne de 8 chiffres CCPPXXXX où
  CC = code catégorie existant, PP = code participant existant (dans cette
  catégorie), XXXX = numéro de test dans la plage attendue pour le canal
  correspondant à l'onglet.
- Toutes les colonnes "Code <n>" ne peuvent contenir que 0, 1, 2,
  « Non applicable » ou « Non observable » (vide autorisé).
- Les colonnes "Code <n> obs"/"...Obs" peuvent contenir n'importe quoi
  (aucun contrôle).
- Id Mystery Tester est obligatoire dès qu'une ligne a un ID Mystery Test.
- Sur l'onglet Phone spécifiquement : Call_Date, Call_hour et Call Duration
  sont également obligatoires (ce sont des colonnes propres à cet onglet).
- Le canal de l'onglet (Phone, Email, Web Navigation, Social Networks, Chat)
  doit être déclaré actif pour le participant visé (case à cocher du canal
  dans Configuration Participant) : un test chargé pour un canal non coché
  chez ce participant est une erreur.

Une ligne entièrement vide est ignorée (pas une erreur).
"""

import re
import datetime as dt

from openpyxl.utils import get_column_letter

# Plages numériques (4 derniers chiffres de l'ID Mystery Test) par canal,
# associées à l'onglet où elles s'appliquent.
CHANNELS = {
    "Phone": {"key": "phone", "range": (1200, 1299)},
    "Email": {"key": "mail", "range": (1300, 1339)},
    "Web Navigation": {"key": "web", "range": (1350, 1364)},
    "Social Networks": {"key": "rs", "range": (1400, 1409)},
    "Chat": {"key": "chat", "range": (1450, 1459)},
}

EXPECTED_SHEETS = list(CHANNELS.keys())

CODE_COLUMN_REGEX = re.compile(r"^code\s*\d+$", re.IGNORECASE)
VALID_CODE_VALUES = {"0", "1", "2", "non applicable", "non observable"}

# Nom du champ Participant (case à cocher) correspondant à chaque canal, pour
# vérifier qu'un test n'est chargé que pour un canal déclaré actif pour ce
# participant dans Configuration Participant.
CHANNEL_FIELD_BY_KEY = {
    "phone": "channel_phone",
    "mail": "channel_mail",
    "web": "channel_web",
    "rs": "channel_rs",
    "chat": "channel_chat",
}

# Colonnes obligatoires supplémentaires, propres à l'onglet Phone.
PHONE_EXTRA_REQUIRED = ["Call_Date", "Call_hour", "Call Duration"]


def _normalize_header(value):
    return re.sub(r"\s+", " ", str(value).strip().lower()) if value else ""


def _find_column(headers, target_name):
    target = _normalize_header(target_name)
    for idx, h in enumerate(headers, start=1):
        if _normalize_header(h) == target:
            return idx
    return None


def _normalize_test_id(raw):
    """Convertit la valeur brute de la cellule ID Mystery Test en chaîne de chiffres."""
    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)
    return str(raw).strip()


def _json_safe(value):
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, dt.timedelta):
        return value.total_seconds()
    return value


def _row_is_empty(row_values):
    return all(v is None or str(v).strip() == "" for v in row_values)


def validate_workbook(wb, categories, participants):
    """
    categories : liste des Category de l'édition en cours
    participants : liste des Participant de l'édition en cours

    Retourne (errors, valid_rows, invalid_channels) :
    - errors : liste de dicts {sheet, row, col_letter, message}
    - valid_rows : liste de dicts prêts à insérer en base (uniquement
      pertinent si errors est vide, puisque le chargement est tout-ou-rien)
    - invalid_channels : dict {nom d'onglet: {noms de participants}} pour les
      tests trouvés sur un canal non déclaré actif chez le participant visé
    """
    category_by_code = {c.code: c for c in categories if c.code}
    participant_by_key = {(p.category_id, p.code): p for p in participants if p.code}

    errors = []
    valid_rows = []
    invalid_channels = {}

    for sheet_name, config in CHANNELS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]

        id_tester_col = _find_column(headers, "Id Mystery Tester")
        extra_required_cols = []
        if sheet_name == "Phone":
            extra_required_cols = [
                (col, _find_column(headers, col)) for col in PHONE_EXTRA_REQUIRED
            ]

        code_columns = [
            (idx, h) for idx, h in enumerate(headers, start=1)
            if h and CODE_COLUMN_REGEX.match(str(h).strip())
        ]

        for r, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_values = list(row_values)
            if _row_is_empty(row_values):
                continue

            row_errors = []
            test_id_str, category, participant = None, None, None

            raw_id = row_values[0] if row_values else None
            if raw_id is None or str(raw_id).strip() == "":
                row_errors.append((1, "ID Mystery Test manquant."))
            else:
                test_id_str = _normalize_test_id(raw_id)
                if len(test_id_str) != 8 or not test_id_str.isdigit():
                    row_errors.append(
                        (1, f"ID Mystery Test invalide ({raw_id!r}) : doit être composé de 8 chiffres (CCPPXXXX).")
                    )
                else:
                    cc, pp, num_str = test_id_str[0:2], test_id_str[2:4], test_id_str[4:8]
                    category = category_by_code.get(cc)
                    if not category:
                        row_errors.append((1, f"Code catégorie « {cc} » inconnu pour cette édition."))
                    else:
                        participant = participant_by_key.get((category.id, pp))
                        if not participant:
                            row_errors.append((1, f"Code participant « {pp} » inconnu pour la catégorie « {cc} »."))
                        else:
                            channel_field = CHANNEL_FIELD_BY_KEY[config["key"]]
                            if not getattr(participant, channel_field):
                                row_errors.append((
                                    1,
                                    f"Le canal « {sheet_name} » n'est pas déclaré actif pour le participant "
                                    f"« {participant.participant_name} » (case décochée dans Configuration Participant).",
                                ))
                                invalid_channels.setdefault(sheet_name, set()).add(participant.participant_name)

                    lo, hi = config["range"]
                    num = int(num_str)
                    if not (lo <= num <= hi):
                        row_errors.append((
                            1,
                            f"Numéro de test {num_str} hors plage pour l'onglet {sheet_name} (attendu {lo}-{hi}).",
                        ))

            if id_tester_col:
                v = row_values[id_tester_col - 1]
                if v is None or str(v).strip() == "":
                    row_errors.append((id_tester_col, "Id Mystery Tester manquant."))

            for col_name, col_idx in extra_required_cols:
                if not col_idx:
                    continue
                v = row_values[col_idx - 1]
                if v is None or str(v).strip() == "":
                    row_errors.append((col_idx, f"{col_name} manquant."))

            for col_idx, header in code_columns:
                v = row_values[col_idx - 1]
                if v is None or str(v).strip() == "":
                    continue
                sval = str(v).strip().lower()
                if sval not in VALID_CODE_VALUES:
                    row_errors.append((
                        col_idx,
                        f"Valeur invalide pour « {header} » : {v!r} (attendu 0, 1, 2, « Non applicable » ou « Non observable »).",
                    ))

            if row_errors:
                for col_idx, message in row_errors:
                    errors.append({
                        "sheet": sheet_name, "row": r,
                        "col_letter": get_column_letter(col_idx), "message": message,
                    })
            else:
                raw_data = {
                    str(headers[i]): _json_safe(row_values[i])
                    for i in range(len(headers)) if headers[i]
                }
                valid_rows.append({
                    "sheet": sheet_name, "channel": config["key"], "row": r,
                    "test_id": test_id_str,
                    "category_id": category.id if category else None,
                    "participant_id": participant.id if participant else None,
                    "raw_data": raw_data,
                })

    return errors, valid_rows, invalid_channels
