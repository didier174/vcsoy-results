"""
Manipulation du fichier Excel « Book scénario » (feuille « step 1 »).

Colonnes de la feuille « step 1 » (A à L) :
  A Validation (0/1 — 1 = ligne validée par l'utilisateur, jamais modifiée)
  B Entreprise (nom du participant)
  C Scénarii (numéro, incrémenté)
  D Prospect/Client
  E Contexte
  F Question
  G Réponse
  H-L : nombre de tests à générer par canal pour ce scénario (voir
  « Générer les tests », app/scenarios/test_generation.py) — jamais
  modifiées par ce module.

La feuille « Recap » n'est jamais lue ni modifiée.
"""

import io

import openpyxl

STEP1_SHEET_NAME = "step 1"

COL_VALIDATION = 1
COL_ENTREPRISE = 2
COL_SCENARII = 3
COL_TYPE = 4
COL_CONTEXTE = 5
COL_QUESTION = 6
COL_REPONSE = 7
# Nombre de tests à générer par canal (voir test_generation.py) :
COL_TEST_PHONE = 8    # H — Téléphone
COL_TEST_MAIL = 9     # I — E-mail
COL_TEST_WEB = 10     # J — Navigation Internet
COL_TEST_RS = 11      # K — Réseaux sociaux
COL_TEST_CHAT = 12    # L — Chat

HEADER_ROW = 1


class BookWorkbookError(Exception):
    pass


def find_step1_sheet(wb):
    if STEP1_SHEET_NAME in wb.sheetnames:
        return wb[STEP1_SHEET_NAME]
    normalized_target = STEP1_SHEET_NAME.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == normalized_target:
            return wb[name]
    raise BookWorkbookError(f"Feuille « {STEP1_SHEET_NAME} » introuvable dans le Book scénario.")


def load_book_state(file_data):
    """
    Lit le workbook et retourne (validated_examples, next_row, next_scenario_num).
    validated_examples : liste de dicts {type, contexte, question, reponse} pour
    les lignes où la colonne A vaut 1 (les plus récentes en dernier).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_data))
    sheet = find_step1_sheet(wb)

    validated_examples = []
    last_row = HEADER_ROW
    last_scenario_num = 0

    row = HEADER_ROW + 1
    while sheet.cell(row=row, column=COL_ENTREPRISE).value not in (None, ""):
        last_row = row
        validation = sheet.cell(row=row, column=COL_VALIDATION).value
        scenario_num = sheet.cell(row=row, column=COL_SCENARII).value
        if isinstance(scenario_num, (int, float)) and scenario_num > last_scenario_num:
            last_scenario_num = int(scenario_num)

        if validation in (1, "1", True):
            validated_examples.append(
                {
                    "type": sheet.cell(row=row, column=COL_TYPE).value or "",
                    "contexte": sheet.cell(row=row, column=COL_CONTEXTE).value or "",
                    "question": sheet.cell(row=row, column=COL_QUESTION).value or "",
                    "reponse": sheet.cell(row=row, column=COL_REPONSE).value or "",
                }
            )
        row += 1

    return validated_examples, last_row + 1, last_scenario_num


def append_scenarios(file_data, participant_name, scenarios):
    """
    Ajoute les scénarios générés (liste de dicts type/contexte/question/reponse)
    à la suite de la feuille « step 1 », sans toucher aux lignes existantes ni
    à la feuille « Recap ». Retourne les bytes du workbook mis à jour, ainsi
    que la liste des numéros de scénario (colonne C) attribués, dans le même
    ordre que `scenarios`.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_data))
    sheet = find_step1_sheet(wb)

    _, next_row, last_scenario_num = load_book_state(file_data)

    assigned_numbers = []
    row = next_row
    scenario_num = last_scenario_num
    for scenario in scenarios:
        scenario_num += 1
        sheet.cell(row=row, column=COL_VALIDATION).value = 0
        sheet.cell(row=row, column=COL_ENTREPRISE).value = participant_name
        sheet.cell(row=row, column=COL_SCENARII).value = scenario_num
        sheet.cell(row=row, column=COL_TYPE).value = scenario.get("type", "")
        sheet.cell(row=row, column=COL_CONTEXTE).value = scenario.get("contexte", "")
        sheet.cell(row=row, column=COL_QUESTION).value = scenario.get("question", "")
        sheet.cell(row=row, column=COL_REPONSE).value = scenario.get("reponse", "")
        assigned_numbers.append(scenario_num)
        row += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), assigned_numbers
