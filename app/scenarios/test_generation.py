"""
Génération du fichier test (« Générer les tests ») à partir du Book
scénario d'un participant : chaque scénario validé est dupliqué en autant
de lignes de test que demandé par canal (colonnes H à L de la feuille
« step 1 », voir excel_utils.py), avec un identifiant de test unique
(catégorie + participant + numéro de plage par canal).

Structure du fichier test généré (feuille active du modèle, colonnes A-P) :
  A Participant       (copié de Entreprise, Book scénario colonne B)
  B Id Test Mystère   (généré : XX catégorie + YY participant + ZZZZ)
  C Scénarii          (copié de Book scénario colonne C)
  D Canal             (valeur fixe selon le canal, voir CHANNEL_SPECS)
  E Prospect/Client   (copié de Book scénario colonne D)
  F, G, H             (jamais touchées)
  I Contexte de la question (copié de Book scénario colonne E)
  J Question          (copié de Book scénario colonne F)
  K Réponse attendue   (copié de Book scénario colonne G)
  L à P               (jamais touchées)
"""

import io

import openpyxl

from app.scenarios import excel_utils

TEST_COL_PARTICIPANT = 1
TEST_COL_ID = 2
TEST_COL_SCENARII = 3
TEST_COL_CANAL = 4
TEST_COL_TYPE = 5
TEST_COL_CONTEXTE = 9
TEST_COL_QUESTION = 10
TEST_COL_REPONSE = 11

HEADER_ROW = 1

# (colonne de comptage dans le Book scénario, libellé du canal, début de plage, fin de plage)
CHANNEL_SPECS = [
    (excel_utils.COL_TEST_PHONE, "Téléphone", 1200, 1299),
    (excel_utils.COL_TEST_MAIL, "E-mail", 1300, 1339),
    (excel_utils.COL_TEST_WEB, "Navigation Internet", 1350, 1364),
    (excel_utils.COL_TEST_RS, "Réseaux sociaux", 1400, 1409),
    (excel_utils.COL_TEST_CHAT, "Chat", 1450, 1459),
]


class TestGenerationError(Exception):
    pass


def validate_test_template_empty(file_data):
    """Lève TestGenerationError si le modèle contient des données à partir
    de la ligne 2 (seule la ligne d'en-tête est acceptée)."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_data))
    except Exception as exc:
        raise TestGenerationError(f"Fichier Excel illisible : {exc}") from exc

    sheet = wb.active
    for row in sheet.iter_rows(min_row=HEADER_ROW + 1, max_row=HEADER_ROW + 1):
        if any(cell.value not in (None, "") for cell in row):
            raise TestGenerationError(
                "Le modèle de fichier test doit contenir uniquement la ligne d'en-tête "
                "(aucune donnée à partir de la ligne 2)."
            )


def _existing_state(sheet):
    """Scanne les lignes déjà présentes dans le fichier test pour connaître
    les (scénario, canal) déjà générés (évite les doublons si on relance) et
    le prochain numéro ZZZZ disponible par canal (continue la numérotation)."""
    processed = set()
    next_zzzz = {label: start for _, label, start, _ in CHANNEL_SPECS}

    row = HEADER_ROW + 1
    while sheet.cell(row=row, column=TEST_COL_PARTICIPANT).value not in (None, ""):
        scenario_num = sheet.cell(row=row, column=TEST_COL_SCENARII).value
        canal = sheet.cell(row=row, column=TEST_COL_CANAL).value
        if scenario_num is not None and canal:
            processed.add((scenario_num, canal))

        test_id = sheet.cell(row=row, column=TEST_COL_ID).value
        if test_id:
            digits = str(test_id)[-4:]
            if digits.isdigit():
                zzzz = int(digits)
                for _, label, start, end in CHANNEL_SPECS:
                    if label == canal and start <= zzzz <= end:
                        next_zzzz[label] = max(next_zzzz[label], zzzz + 1)
        row += 1

    return processed, next_zzzz, row


def generate_test_rows(test_file_data, book_file_data, participant_name, category_code, participant_code):
    """
    Ajoute au fichier test (copie du modèle ou fichier existant) les lignes
    de test correspondant aux scénarios du Book scénario, en dupliquant
    chaque scénario selon le nombre indiqué par canal (colonnes H-L).
    Les couples (scénario, canal) déjà présents dans le fichier test ne
    sont pas régénérés (relance sans risque de doublon).

    Retourne (nouveaux bytes du fichier test, nombre de lignes ajoutées).
    """
    book_wb = openpyxl.load_workbook(io.BytesIO(book_file_data))
    book_sheet = excel_utils.find_step1_sheet(book_wb)

    test_wb = openpyxl.load_workbook(io.BytesIO(test_file_data))
    test_sheet = test_wb.active

    processed, next_zzzz, write_row = _existing_state(test_sheet)

    added = 0
    book_row = excel_utils.HEADER_ROW + 1
    while book_sheet.cell(row=book_row, column=excel_utils.COL_ENTREPRISE).value not in (None, ""):
        scenario_num = book_sheet.cell(row=book_row, column=excel_utils.COL_SCENARII).value
        type_value = book_sheet.cell(row=book_row, column=excel_utils.COL_TYPE).value
        contexte_value = book_sheet.cell(row=book_row, column=excel_utils.COL_CONTEXTE).value
        question_value = book_sheet.cell(row=book_row, column=excel_utils.COL_QUESTION).value
        reponse_value = book_sheet.cell(row=book_row, column=excel_utils.COL_REPONSE).value

        for count_col, channel_label, range_start, range_end in CHANNEL_SPECS:
            count = book_sheet.cell(row=book_row, column=count_col).value
            if not isinstance(count, (int, float)) or count <= 0:
                continue
            if (scenario_num, channel_label) in processed:
                continue

            for _ in range(int(count)):
                zzzz = next_zzzz[channel_label]
                if zzzz > range_end:
                    raise TestGenerationError(
                        f"Plage de numéros de test épuisée pour le canal « {channel_label} » "
                        f"({range_start}-{range_end})."
                    )
                test_id = f"{category_code}{participant_code}{zzzz:04d}"

                test_sheet.cell(row=write_row, column=TEST_COL_PARTICIPANT).value = participant_name
                test_sheet.cell(row=write_row, column=TEST_COL_ID).value = test_id
                test_sheet.cell(row=write_row, column=TEST_COL_SCENARII).value = scenario_num
                test_sheet.cell(row=write_row, column=TEST_COL_CANAL).value = channel_label
                test_sheet.cell(row=write_row, column=TEST_COL_TYPE).value = type_value
                test_sheet.cell(row=write_row, column=TEST_COL_CONTEXTE).value = contexte_value
                test_sheet.cell(row=write_row, column=TEST_COL_QUESTION).value = question_value
                test_sheet.cell(row=write_row, column=TEST_COL_REPONSE).value = reponse_value

                next_zzzz[channel_label] = zzzz + 1
                write_row += 1
                added += 1

            processed.add((scenario_num, channel_label))

        book_row += 1

    buffer = io.BytesIO()
    test_wb.save(buffer)
    return buffer.getvalue(), added
